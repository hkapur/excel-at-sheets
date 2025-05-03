import os
import re
import pandas as pd
import matplotlib
matplotlib.use('Agg') # Use non-interactive backend for server use
import matplotlib.pyplot as plt
import uuid
from pathlib import Path # To create directory
import base64 # Import base64
from typing import Literal, Optional # For Pydantic model
from pydantic import BaseModel, Field # For Pydantic model
import fitz  # PyMuPDF
import pytesseract
from PIL import Image # To handle images for OCR
from langchain.output_parsers import PydanticOutputParser

from openpyxl import load_workbook
from langchain_openai import ChatOpenAI
from langchain.schema import HumanMessage, SystemMessage
from markdown import markdown
from weasyprint import HTML, CSS
from weasyprint.text.fonts import FontConfiguration
import io

# Initialize LLM
llm = ChatOpenAI(model="gpt-4o-mini", temperature=0.3, api_key=os.getenv("OPENAI_API_KEY"))

# Ensure the directory for saving charts exists
CHARTS_DIR = Path("static") / "generated_charts"
CHARTS_DIR.mkdir(parents=True, exist_ok=True)

# ========= Pydantic Model for Tool Arguments ==========
class ChartRequestArgs(BaseModel):
    """Schema for arguments of the chart request tool."""
    chart_type: Literal["bar", "histogram", "scatter", "none"] = Field(description="The type of chart requested (bar, histogram, scatter, or none).")
    column1: Optional[str] = Field(None, description="The primary column name for the chart (required if chart_type is not 'none').")
    column2: Optional[str] = Field(None, description="The secondary column name (required only if chart_type is 'scatter').")

# ========= Pydantic Model for Receipt Data ==========
class ReceiptDetails(BaseModel):
    """Schema for extracted receipt details."""
    vendor: Optional[str] = Field(None, description="The name of the vendor or store.")
    date: Optional[str] = Field(None, description="The date of the transaction (e.g., YYYY-MM-DD or MM/DD/YYYY).")
    total: Optional[str] = Field(None, description="The final total amount paid.")

class ReceiptList(BaseModel):
    """Schema for a list of extracted receipts."""
    receipts: list[ReceiptDetails] = Field(description="A list of all receipts found in the text.")

# ========= LangChain Helpers =========

def ask_about_spreadsheet(table_text, user_question):
    system = "You are a helpful assistant that analyzes spreadsheets with multiple sheets and answers clearly."
    user = f"""
Use bullet points and markdown formatting in your response.

**Spreadsheet content:**

{table_text}

**User question:** {user_question}
"""
    try:
        response = llm.invoke([SystemMessage(content=system), HumanMessage(content=user)])
        return response.content
    except Exception as e:
        return f"Error: {str(e)}"

def explain_formula(formula):
    system = "You are a helpful assistant who explains Excel formulas step-by-step using markdown."
    user = f"Explain this Excel formula:\n\n`{formula}`"
    try:
        response = llm.invoke([SystemMessage(content=system), HumanMessage(content=user)])
        return response.content
    except Exception as e:
        return f"Error: {str(e)}"

def get_column_summary(df):
    """Creates a text summary of DataFrame columns for the LLM, including basic stats for numeric columns."""
    if df is None or df.empty:
        return "The spreadsheet is empty or could not be read."
    
    summary_lines = ["Spreadsheet Columns:"]
    for col in df.columns:
        dtype = str(df[col].dtype)
        line = f"- '{col}' (Type: {dtype})"
        
        # Add unique count for potential categorical columns
        if dtype in ['object', 'category']:
            # Also check for boolean type which might be treated as object sometimes
            is_bool_like = df[col].nunique() == 2 and df[col].dropna().isin([0, 1, True, False]).all()
            if is_bool_like:
                 line += f" (Boolean-like: {df[col].unique().tolist()})" # Show the two values
            else:
                num_unique = df[col].nunique()
                line += f" ({num_unique} unique values)"
                # Show some sample values if cardinality is low
                if num_unique < 10:
                    sample_values = df[col].dropna().unique()[:5]
                    line += f" Samples: [{ ', '.join(map(str, sample_values)) }]"
                    
        # Add basic stats for numeric columns
        elif pd.api.types.is_numeric_dtype(df[col]):
            stats = df[col].describe()
            min_val = stats.get('min', 'N/A')
            max_val = stats.get('max', 'N/A')
            mean_val = stats.get('mean', 'N/A')
            line += f" (Min: {min_val:.2f}, Max: {max_val:.2f}, Mean: {mean_val:.2f})" if pd.notna(min_val) else " (Numeric)"
            
        summary_lines.append(line)
        
    print("--- Generated Column Summary for LLM --- \n" + "\n".join(summary_lines) + "\n-------------------------------------")
    return "\n".join(summary_lines)

def generate_analysis_report(df, table_text):
    """Generates ONLY the text analysis report using the LLM."""
    
    column_summary = get_column_summary(df)
    
    # Prompt focused ONLY on generating the text report
    system = f'''You are a professional data analyst assistant. Analyze the column summary and data sample provided below.

**Task:** Generate a comprehensive text-based analysis report.

**Important Note:** In a subsequent step, you will be asked (via a tool call) to specify the most appropriate chart (bar, histogram, scatter, or none) based on this data. Please write your analysis text *anticipating* the chart you intend to request. For example, if you plan to request a bar chart of 'Category', your 'Key Findings' should include insights derivable from such a chart.

**Report Structure Requirements:**
1. Objective and Scope
2. Data Description
3. Methodology (Describe analysis based on data sample)
4. Visualizations and Tables (Briefly describe the chart you *intend* to request later, or state why none is appropriate)
5. Key Findings (Derived ONLY from the provided data sample, *incorporating insights from the anticipated chart*)
6. Conclusions and Recommendations (Based ONLY on the provided data sample *and anticipated chart insights*)
7. Professional Formatting (Use markdown headings, bullet points, etc.)

Generate *only* the narrative report content. Do NOT include placeholders for charts or the actual tool call request in this text.'''

    user = f'''
**Column Summary:**
{column_summary}

**Spreadsheet Data Sample:**
{table_text}

**Generate the text analysis report based ONLY on the summary and sample provided.**
'''

    try:
        # Invoke LLM for text generation ONLY
        response = llm.invoke([SystemMessage(content=system), HumanMessage(content=user)])
        report_text = response.content
        print("--- LLM Text Report Generation Complete ---")
        return report_text.strip()

    except Exception as e:
        print(f"Error in generate_analysis_report (text only): {e}")
        import traceback
        traceback.print_exc() 
        return f"Error generating text report: {str(e)}"

# ========= New Function for Chart Decision ==========
def decide_chart_request(df):
    """Asks the LLM to decide on a chart via tool call based on data structure."""
    if df is None or df.empty:
        print("Chart Decision: DataFrame empty, skipping LLM call.")
        return None
        
    column_summary = get_column_summary(df)

    # Refined prompt
    system = f'''You are a data visualization expert. Based on the column summary, decide if a **single meaningful chart** (bar, histogram, scatter) can effectively visualize a key aspect or distribution within the data.

**Guidelines:**
- Use `histogram` for visualizing the distribution of a numerical column (like age, duration, price).
- Use `bar` for showing counts of distinct categories or boolean values (like status, type, escalated/not escalated).
- Use `scatter` *only* if there are two numerical columns that might show a correlation.
- If no single chart provides significant insight or if data types are unsuitable, choose 'none'.

**Instructions:**
1. Analyze the columns, types, stats, and unique value counts.
2. **CRITICAL:** Use the `register_chart_request` tool to register your decision. Provide `chart_type`, `column1`, and `column2` (if applicable).
   - For `bar` or `histogram`, specify `column1`.
   - For `scatter`, specify `column1` and `column2`.
   - For `none`, set `chart_type` to `none`.
Do NOT generate any other text output, only make the tool call.'''

    user = f'''
**Column Summary:**
{column_summary}

**Use the `register_chart_request` tool to specify the best chart based on this summary, or 'none' if no single chart is clearly insightful.**
'''

    try:
        llm_with_tool = llm.bind_tools([ChartRequestArgs])
        response = llm_with_tool.invoke([SystemMessage(content=system), HumanMessage(content=user)])

        # Process tool calls
        tool_args = None
        tool_calls = getattr(response, 'tool_calls', [])
        if tool_calls:
            first_tool_call = tool_calls[0]
            if first_tool_call.get('name') == ChartRequestArgs.__name__:
                 tool_args = first_tool_call.get('args')
                 print(f"Chart Decision: Parsed Tool Call Args: {tool_args}")
            else:
                 print(f"Chart Decision Warning: Unexpected tool call: {first_tool_call.get('name')}")
        else:
            # Check additional_kwargs as fallback
            additional_kwargs = getattr(response, 'additional_kwargs', {})
            tool_call_data = additional_kwargs.get('tool_calls')
            if tool_call_data and isinstance(tool_call_data, list) and len(tool_call_data) > 0:
                 first_tool_call_legacy = tool_call_data[0]
                 if first_tool_call_legacy.get('function', {}).get('name') == ChartRequestArgs.__name__:
                    try:
                        import json
                        tool_args = json.loads(first_tool_call_legacy.get('function', {}).get('arguments', '{}'))
                        print(f"Chart Decision: Parsed Tool Call Args (from additional_kwargs): {tool_args}")
                    except json.JSONDecodeError:
                        print("Chart Decision Warning: Failed to parse tool call args from additional_kwargs.")
                 else:
                     print(f"Chart Decision Warning: Unexpected function name in additional_kwargs: {first_tool_call_legacy.get('function', {}).get('name')}")
            else:
                print("Chart Decision Warning: LLM did not make the expected tool call.")

        return tool_args # Return the dictionary of args or None

    except Exception as e:
        print(f"Error in decide_chart_request: {e}")
        import traceback
        traceback.print_exc()
        return None # Return None on error

# ========= PDF Generation Helper =========
def generate_report_pdf(markdown_content, chart_image_abs_path=None):
    """Converts Markdown report content (including chart reference) to PDF bytes."""
    
    html_content = markdown(markdown_content, extensions=["fenced_code", "tables"])
    
    # --- Image Embedding Logic --- 
    if chart_image_abs_path and os.path.exists(chart_image_abs_path):
        try:
            # 1. Read image data
            with open(chart_image_abs_path, "rb") as image_file:
                image_data = image_file.read()
            
            # 2. Encode as Base64
            base64_encoded_data = base64.b64encode(image_data).decode('utf-8')
            
            # 3. Create data URI (assuming PNG, adjust if other formats are possible)
            mime_type = "image/png" 
            data_uri = f"data:{mime_type};base64,{base64_encoded_data}"
            print(f"PDF Generation: Created data URI (length: {len(data_uri)})")

            # 4. Find the relative path pattern in the original markdown content
            chart_path_match = re.search(r'!\[.*?\]\((/static/generated_charts/.*?)\)', markdown_content)
            
            if chart_path_match:
                chart_relative_path = chart_path_match.group(1)
                print(f"PDF Generation: Found relative path in markdown: {chart_relative_path}")
                
                # 5. Find and replace the <img> tag's src in the generated HTML
                img_tag_pattern = f'<img [^>]*src="{re.escape(chart_relative_path)}"[^>]*>'
                # More flexible replacement focusing only on the src attribute
                def replace_src(match):
                    return match.group(0).replace(f'src="{chart_relative_path}"' , f'src="{data_uri}"')

                html_content_new, replacements_made = re.subn(
                    img_tag_pattern, 
                    replace_src,
                    html_content,
                    count=1 
                )
                
                if replacements_made > 0:
                    html_content = html_content_new
                    print(f"PDF Generation: Replaced relative src '{chart_relative_path}' with data URI.")
                else:
                    print(f"PDF Generation Warning: Could not find matching <img> tag src for relative path '{chart_relative_path}' in HTML content.")
                    # Fallback: try replacing raw string - less reliable but might catch cases
                    # where markdown parser changed attributes around src
                    if f'src="{chart_relative_path}"' in html_content:
                        html_content = html_content.replace(f'src="{chart_relative_path}"' , f'src="{data_uri}"' , 1)
                        print(f"PDF Generation: (Fallback) Replaced raw src string with data URI.")
                    else:
                        print(f"PDF Generation Warning: Fallback replacement also failed.")
            else:
                print("PDF Generation Warning: Could not find chart path pattern in markdown content.")
        except Exception as e:
            print(f"PDF Generation Error during image embedding: {e}")
            # Continue without the image if embedding fails

    elif chart_image_abs_path:
        print(f"PDF Generation Warning: Provided chart path does not exist: {chart_image_abs_path}")
    else:
         print("PDF Generation: No chart path provided for embedding.")
    # --- End Image Embedding Logic ---

    # Basic CSS for PDF styling
    css = CSS(string='''
        @page { size: A4; margin: 2cm; }
        body { font-family: sans-serif; line-height: 1.4; }
        h1, h2, h3, h4, h5, h6 { font-weight: bold; margin-top: 1.5em; margin-bottom: 0.5em; }
        h1 { font-size: 1.8em; }
        h2 { font-size: 1.5em; }
        h3 { font-size: 1.2em; }
        p { margin-bottom: 1em; }
        ul, ol { margin-left: 20px; margin-bottom: 1em; }
        li { margin-bottom: 0.5em; }
        table { border-collapse: collapse; width: 100%; margin-bottom: 1em; }
        th, td { border: 1px solid #ccc; padding: 8px; text-align: left; }
        th { background-color: #f2f2f2; }
        img { max-width: 100%; height: auto; display: block; margin: 1em auto; }
        pre { background-color: #f5f5f5; padding: 10px; border-radius: 4px; overflow-x: auto; }
        code { font-family: monospace; }
    ''')
    
    font_config = FontConfiguration()

    # Generate PDF bytes using the modified HTML
    try:
        html = HTML(string=html_content)
        pdf_bytes = html.write_pdf(stylesheets=[css], font_config=font_config)
        print("PDF Generation: PDF bytes created successfully.")
        return pdf_bytes
    except Exception as e:
        print(f"Error generating PDF with WeasyPrint: {e}")
        return None

# ========= Spreadsheet Helpers =========

def extract_table_data_all_sheets(file_path, max_rows=30, max_cols=10):
    """Extracts data from the first sheet into a pandas DataFrame 
       and also creates a text representation of all sheets (limited rows/cols)."""
    
    # Read the first sheet into a pandas DataFrame
    try:
        df = pd.read_excel(file_path, sheet_name=0) # Read first sheet (index 0)
    except Exception as e:
        print(f"Error reading Excel file into DataFrame: {e}")
        df = pd.DataFrame() # Return empty DataFrame on error

    # Create limited text representation (existing logic)
    all_data = []    
    try:
        wb = load_workbook(file_path) # Need to reload for openpyxl processing
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            all_data.append(f"### Sheet: {sheet}")
            # Limit rows/cols for text preview
            preview_rows = min(ws.max_row, max_rows)
            preview_cols = min(ws.max_column, max_cols)
            for row in ws.iter_rows(min_row=1, max_row=preview_rows, max_col=preview_cols, values_only=True):
                row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
                all_data.append(row_text)
            all_data.append("")  # spacing
    except Exception as e:
         print(f"Error processing sheets with openpyxl: {e}")
         all_data.append("[Error processing sheet data for text preview]")

    return df, "\n".join(all_data)

def get_cell_value_across_sheets(wb, cell):
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        try:
            val = ws[cell].value
            if val is not None:
                return val
        except:
            continue
    return "(not found in any sheet)"

def build_formula_chain(wb, formula):
    refs = re.findall(r'\b[A-Z]{1,2}[0-9]{1,4}\b', formula)
    if not refs:
        return ""
    parts = ["\n**Referenced cells:**"]
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for ref in refs:
            try:
                val = ws[ref].value
                if val is not None:
                    parts.append(f"- `{ref}` in *{sheet}* = `{val}`")
            except:
                continue
    return "\n".join(parts)

def generate_dynamic_prompts(df, max_prompts=3):
    """Generates a list of relevant prompt suggestions based on DataFrame columns."""
    prompts = []
    if df is None or df.empty:
        return prompts

    categorical_cols = df.select_dtypes(include=['object', 'category']).columns
    numerical_cols = df.select_dtypes(include=['number']).columns

    # Suggestion 1: Value counts for first categorical column
    if len(categorical_cols) > 0:
        col_name = categorical_cols[0]
        # Check if column has a reasonable number of unique values for counts
        if df[col_name].nunique() < 50: # Avoid suggesting counts for ID-like columns
             prompts.append(f"Show value counts for '{col_name}'.")
    
    # Suggestion 2: Average for first numerical column
    if len(numerical_cols) > 0:
        prompts.append(f"What is the average of '{numerical_cols[0]}'?")

    # Suggestion 3: Unique values for second categorical column (if exists)
    if len(categorical_cols) > 1:
         col_name = categorical_cols[1]
         if df[col_name].nunique() < 50:
            prompts.append(f"What are the unique values in '{categorical_cols[1]}'?")
    # If no second categorical, try sum of first numerical (if not already added)
    elif len(numerical_cols) > 0 and len(prompts) < max_prompts:
         prompts.append(f"What is the sum of '{numerical_cols[0]}'?")

    # Add more generic prompts if we still have space
    if len(prompts) < max_prompts:
        prompts.append("Summarize the first few rows.")
    if len(prompts) < max_prompts and len(df.columns) > 0:
         prompts.append(f"What are the column names?")

    return prompts[:max_prompts]

# ========= Utility Helpers =========

# --- PDF/OCR Function ---
def process_pdf_with_ocr(pdf_path):
    """Extracts text from all pages of a PDF using PyMuPDF and Tesseract OCR."""
    all_text = ""
    print(f"Starting OCR process for: {pdf_path}")
    try:
        doc = fitz.open(pdf_path)
        print(f"Opened PDF: {pdf_path}, Pages: {len(doc)}")
        
        for page_num in range(len(doc)):
            page = doc.load_page(page_num)
            print(f"Processing Page {page_num + 1}/{len(doc)}...")
            
            # Render page to an image (pixmap)
            # Increase DPI for better OCR results if needed, e.g., dpi=300
            pix = page.get_pixmap(dpi=200) 
            
            # Convert pixmap to PIL Image
            img_bytes = pix.tobytes("png") # Use PNG format
            img = Image.open(io.BytesIO(img_bytes))
            
            # Perform OCR using pytesseract
            try:
                page_text = pytesseract.image_to_string(img)
                all_text += page_text + "\n\n---\n\n" # Add separator between pages
                print(f"  Extracted ~{len(page_text)} characters from page {page_num + 1}")
            except pytesseract.TesseractNotFoundError:
                 print("ERROR: Tesseract executable not found. Please install Tesseract OCR.")
                 raise # Re-raise the specific error
            except Exception as ocr_err:
                 print(f"  Warning: Pytesseract error on page {page_num + 1}: {ocr_err}")
                 # Optionally append an error marker to the text
                 all_text += f"[OCR Error on Page {page_num + 1}]\n\n---\n\n"
                 
        doc.close()
        print(f"Finished OCR. Total text length: {len(all_text)}")
        return all_text.strip()
        
    except fitz.fitz.FileNotFoundError:
         print(f"Error: PDF file not found at {pdf_path}")
         return "Error: PDF file not found."
    except pytesseract.TesseractNotFoundError: # Catch again if raised from loop
         return "Error: Tesseract OCR engine not installed or not found in PATH."
    except Exception as e:
        print(f"Error opening or processing PDF {pdf_path}: {e}")
        import traceback
        traceback.print_exc()
        return f"Error processing PDF: {str(e)}"

# --- LLM Receipt Parser Function ---
def extract_receipt_data_llm(ocr_text):
    """Uses LLM with PydanticOutputParser to extract structured data for multiple receipts from OCR text."""
    print("Starting LLM multiple receipt data extraction...")
    
    if not ocr_text or ocr_text.startswith("Error:"):
        print(f"Skipping LLM extraction due to invalid OCR text: {ocr_text}")
        return pd.DataFrame([{"vendor": "OCR Error", "date": None, "total": None}])
        
    # 1. Initialize Parser with the LIST model
    parser = PydanticOutputParser(pydantic_object=ReceiptList) # Use ReceiptList
    
    # 2. Define Prompt for MULTIPLE receipts
    format_instructions = parser.get_format_instructions()
    system_prompt = f"""You are an expert assistant specialized in extracting information from OCR text containing one or more receipts, potentially across multiple pages (indicated by '---' separators).
Extract the vendor name, transaction date, and total amount for **each distinct receipt** found in the text. 
Return the results as a list of objects. If a field is not found for a specific receipt, leave it null for that receipt.
{format_instructions}
"""
    user_prompt = f"Receipt Text (potentially multiple receipts):\n```\n{ocr_text}\n```\nPlease extract the information for all receipts found."

    # 3. Invoke LLM
    print("Invoking LLM for structured extraction of multiple receipts...")
    try:
        messages = [
            SystemMessage(content=system_prompt),
            HumanMessage(content=user_prompt)
        ]
        response = llm.invoke(messages)
        llm_output = response.content
        print(f"LLM Raw Output:\n{llm_output}")

        # 4. Parse Response (expecting ReceiptList)
        try:
            parsed_data = parser.parse(llm_output)
            print(f"Successfully parsed LLM list response: Found {len(parsed_data.receipts)} receipts")
            
            # 5. Convert list of Pydantic objects to DataFrame
            if parsed_data.receipts:
                # Convert each ReceiptDetails object in the list to a dict
                list_of_dicts = [receipt.dict() for receipt in parsed_data.receipts]
                # --- Add Logging Here ---
                print("--- Parsed Receipt Dicts (before DataFrame) ---")
                import pprint
                pprint.pprint(list_of_dicts)
                print("---------------------------------------------")
                # --- End Logging ---
                df = pd.DataFrame(list_of_dicts)
                print("Converted list of parsed receipts to DataFrame.")
            else:
                 print("LLM parsed successfully but returned an empty list of receipts.")
                 df = pd.DataFrame(columns=['vendor', 'date', 'total']) # Return empty DataFrame
            return df
            
        except Exception as parse_error: # Catch parsing errors specifically
            print(f"Error parsing LLM list response: {parse_error}")
            print("LLM Output that failed parsing:")
            print(llm_output) 
            return pd.DataFrame([{
                "vendor": "LLM Parse Error", 
                "date": None, 
                "total": f"Parse Error: {str(parse_error)[:100]}..."
            }])

    except Exception as llm_error:
        print(f"Error during LLM invocation for multi-receipt parsing: {llm_error}")
        import traceback
        traceback.print_exc()
        return pd.DataFrame([{
            "vendor": "LLM Invocation Error", 
            "date": None, 
            "total": f"LLM Error: {str(llm_error)[:100]}..."
        }])
# --- End LLM Receipt Parser ---

def extract_cell(text):
    match = re.search(r"\b([A-Z]{1,2}[0-9]{1,4})\b", text.upper())
    return match.group(1) if match else "A1"

def is_cell_reference(text):
    return bool(re.search(r"\b([A-Z]{1,2}[0-9]{1,4})\b", text.upper()))

# ========= Helper functions for plotting (add these) =========
def _plot_bar(df, col_name, chart_path):
    """Generates and saves a bar chart."""
    try:
        if col_name not in df.columns:
            print(f"Plot Error: Column '{col_name}' not found for bar chart.")
            return False
        if df[col_name].nunique() > 50 or df[col_name].nunique() < 2:
            print(f"Plot Error: Column '{col_name}' has unsuitable number of unique values ({df[col_name].nunique()}) for bar chart.")
            return False
        
        plt.figure(figsize=(10, 6))
        df[col_name].value_counts().plot(kind='bar')
        plt.title(f'Counts by {col_name}')
        plt.xlabel(col_name)
        plt.ylabel('Count')
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        plt.savefig(chart_path)
        plt.close()
        print(f"Plot Success: Saved bar chart for '{col_name}' to {chart_path}")
        return True
    except Exception as e:
        print(f"Plot Error (Bar): {e}")
        return False

def _plot_histogram(df, col_name, chart_path):
    """Generates and saves a histogram."""
    try:
        if col_name not in df.columns:
            print(f"Plot Error: Column '{col_name}' not found for histogram.")
            return False
        if not pd.api.types.is_numeric_dtype(df[col_name]):
            print(f"Plot Error: Column '{col_name}' is not numeric for histogram.")
            return False
        if df[col_name].nunique() <= 1:
            print(f"Plot Error: Column '{col_name}' needs more than 1 unique value for histogram.")
            return False
        
        plt.figure(figsize=(10, 6))
        df[col_name].plot(kind='hist', bins=15)
        plt.title(f'Distribution of {col_name}')
        plt.xlabel(col_name)
        plt.ylabel('Frequency')
        plt.tight_layout()
        plt.savefig(chart_path)
        plt.close()
        print(f"Plot Success: Saved histogram for '{col_name}' to {chart_path}")
        return True
    except Exception as e:
        print(f"Plot Error (Histogram): {e}")
        return False

def _plot_scatter(df, col1_name, col2_name, chart_path):
    """Generates and saves a scatter plot."""
    try:
        if col1_name not in df.columns or col2_name not in df.columns:
            print(f"Plot Error: One or both columns ('{col1_name}', '{col2_name}') not found for scatter.")
            return False
        if not pd.api.types.is_numeric_dtype(df[col1_name]) or not pd.api.types.is_numeric_dtype(df[col2_name]):
            print(f"Plot Error: Both columns ('{col1_name}', '{col2_name}') must be numeric for scatter.")
            return False
            
        plt.figure(figsize=(10, 6))
        plt.scatter(df[col1_name], df[col2_name])
        plt.title(f'Scatter Plot of {col1_name} vs {col2_name}')
        plt.xlabel(col1_name)
        plt.ylabel(col2_name)
        plt.grid(True)
        plt.tight_layout()
        plt.savefig(chart_path)
        plt.close()
        print(f"Plot Success: Saved scatter plot for '{col1_name}' vs '{col2_name}' to {chart_path}")
        return True
    except Exception as e:
        print(f"Plot Error (Scatter): {e}")
        return False

# ========= Receipt Analysis Helpers ==========

def categorize_receipts(receipts_with_std_date):
    """Uses LLM to assign a category to each receipt based on vendor.
       Expects input dicts to have 'vendor', 'total', and 'date' (standardized).
       Returns a list of dicts including 'vendor', 'total', 'date', and 'category'."""
    if not receipts_with_std_date:
        return []
        
    system = "You are an expense categorization assistant." 
    # Updated prompt to specify expected input/output including date
    user = "Categorize each receipt into a spending category. Common categories: Food, Transport, Shopping, Utilities, Entertainment, Other.\n"
    user += "Each item has Vendor, Total, and Date.\n"
    user += "Return the result as JSON list with: vendor, total, date, category.\n\n"

    for r in receipts_with_std_date:
        # Include date in the info passed to LLM, although categorization is mainly by vendor
        user += f"- Vendor: {r.get('vendor', 'N/A')}, Total: {r.get('total', 'N/A')}, Date: {r.get('date', 'N/A')}\n"

    # Updated example output to include date
    example = """
Example Output:
[
  {"vendor": "Starbucks", "total": 12.50, "date": "2024-10-20", "category": "Food"},
  {"vendor": "Amazon", "total": 59.99, "date": "2024-10-21", "category": "Shopping"}
]
"""
    print(f"Sending {len(receipts_with_std_date)} receipts to LLM for categorization...")
    try:
        from langchain.schema import SystemMessage, HumanMessage
        response = llm.invoke([SystemMessage(content=system), HumanMessage(content=user + example)])

        print("--- LLM raw categorization output ---")
        print(repr(response.content))
        print("-----------------------------------")

        # Strip Markdown code block
        cleaned = response.content.strip()
        if cleaned.startswith("```json"):
            cleaned = cleaned[len("```json"):].strip()
        if cleaned.endswith("```"):
            cleaned = cleaned[:-3].strip()
        
        # Attempt to repair potentially truncated JSON
        if not cleaned.endswith("]"):
            if cleaned.rfind('}') > cleaned.rfind('{'): # Simple check for dangling comma maybe
                cleaned = cleaned[:cleaned.rfind('}')+1] + "]"
                print("Attempted basic JSON repair (closing bracket).")
            else: 
                 print("Warning: JSON from LLM might be incomplete.")

        import json
        categorized_list = json.loads(cleaned)
        print(f"Successfully parsed {len(categorized_list)} categorized items from LLM.")
        return categorized_list
    except json.JSONDecodeError as json_err:
        print(f"Error decoding JSON from LLM categorization: {json_err}")
        print(f"Failed JSON string: {cleaned}")
        # Fallback: return original list but add an 'Error' category?
        for r in receipts_with_std_date:
            r['category'] = "Categorization Failed"
        return receipts_with_std_date 
    except Exception as e:
        print(f"Error categorizing receipts: {e}")
        # Fallback
        for r in receipts_with_std_date:
            r['category'] = "Categorization Error"
        return receipts_with_std_date

def create_transaction_table(categorized_data):
    """Creates a markdown table from the categorized receipt data."""
    if not categorized_data:
        return "No transaction data available to create table."

    # Convert list of dicts to DataFrame
    df = pd.DataFrame(categorized_data)
    
    # Ensure required columns exist, add if missing
    required_cols = ['vendor', 'date', 'category', 'total']
    for col in required_cols:
        if col not in df.columns:
            df[col] = None # Or some default value
            
    # Select and rename columns for the final table
    df_table = df[['vendor', 'date', 'category', 'total']].copy()
    df_table.rename(columns={
        'vendor': 'Vendor',
        'date': 'Date', 
        'category': 'Category',
        'total': 'Amount ($)'
    }, inplace=True)
    
    # Format the Amount column potentially
    # df_table['Amount ($)'] = pd.to_numeric(df_table['Amount ($)'], errors='coerce').fillna(0).map('{:.2f}'.format)

    print(f"Creating markdown table for {len(df_table)} transactions.")
    try:
        # Generate markdown table, aligning numeric column right
        # Note: Requires pandas >= 1.0.0 for floatfmt
        # Note: Requires tabulate package to be installed
        markdown_table = df_table.to_markdown(index=False, floatfmt=".2f") 
        return markdown_table
    except Exception as e:
        print(f"Error generating markdown table: {e}")
        return "Error generating transaction table."

def plot_expense_pie_chart(categorized_data, save_path):
    """Generates a pie chart of expenses grouped by category, showing amount and percentage."""
    import pandas as pd
    import matplotlib.pyplot as plt

    if not categorized_data:
        print("Plot Error: No categorized data provided for pie chart.")
        return False

    try:
        df = pd.DataFrame(categorized_data)
        # Ensure total is numeric
        df['total'] = pd.to_numeric(df['total'], errors='coerce')
        df.dropna(subset=['total'], inplace=True)

        if df.empty:
            print("Plot Error: No valid numeric expense data found after cleaning.")
            return False
            
        grouped = df.groupby("category")["total"].sum()
        total = grouped.sum() # Calculate total for label formatting
        
        # Custom autopct function to display value and percentage
        def make_autopct(values):
            def my_autopct(pct):
                # --- Debugging Removed --- 
                if pct is None or total is None or total == 0:
                     return '' # Avoid division by zero or errors with None
                try:
                    absolute = (float(pct) / 100.0) * float(total)
                    # Use SPACE instead of NEWLINE in format string
                    formatted_string = f"${absolute:.2f} ({pct:.1f}%)" 
                    # --- Debugging Removed --- 
                    return formatted_string
                except (ValueError, TypeError) as e:
                     print(f"    Error formatting autopct label: {e}. Pct: {pct}")
                     # Fallback to just percentage on error
                     try:
                         return f'{float(pct):.1f}%' 
                     except:
                         return '' # Give up if pct is totally invalid
            return my_autopct

        plt.figure(figsize=(8, 8))
        # Use the custom function for autopct
        grouped.plot.pie(autopct=make_autopct(grouped.values), startangle=90, pctdistance=0.85) 
        plt.title("Expenses by Category", pad=20) # Added padding to title
        plt.ylabel("") # Keep Y label empty for pie charts
        # plt.tight_layout() # Can sometimes interfere with pie labels, test if needed
        plt.savefig(save_path)
        plt.close()
        print(f"Pie chart with amount and percentage saved to {save_path}")
        return True
    except Exception as e:
        print(f"Error generating pie chart: {e}")
        import traceback
        traceback.print_exc()
        return False

def generate_expense_summary(categorized_data):
    """LLM generates a summary of user's expenses by category."""
    import pandas as pd
    df = pd.DataFrame(categorized_data)
    category_totals = df.groupby("category")["total"].sum().to_dict()

    user = f"""
Below are total expenses per category:

{category_totals}

Write a short summary of the user's spending pattern. Be concise and professional.
"""
    try:
        from langchain.schema import SystemMessage, HumanMessage
        response = llm.invoke([SystemMessage(content="You are an expense report assistant."), HumanMessage(content=user)])
        return response.content.strip()
    except Exception as e:
        print(f"Error generating summary: {e}")
        return "Error generating summary."

def generate_financial_analysis(categorized_data, period_str, total_spent_str, num_transactions):
    """LLM generates a structured financial analysis with summary, observations, and suggestions."""
    import pandas as pd
    print("Generating structured financial analysis...")
    
    if not categorized_data:
        return "Could not generate analysis: No categorized data provided."
        
    try:
        df = pd.DataFrame(categorized_data)
        df['total'] = pd.to_numeric(df['total'], errors='coerce')
        df.dropna(subset=['total'], inplace=True)

        if df.empty:
            return "Could not generate analysis: No valid expense data found after processing."

        # Calculate category totals and percentages for the prompt
        category_summary = df.groupby("category")["total"].sum()
        total_calculated = category_summary.sum() # Use sum from grouped data for consistency
        category_percentages = (category_summary / total_calculated * 100).round(1)
        
        summary_lines = []
        for category, total in category_summary.items():
            percentage = category_percentages.get(category, 0)
            summary_lines.append(f"{category}: ${total:.2f} ({percentage}%)")
        category_summary_text = "\n".join(summary_lines)

        # --- New Structured Prompt ---
        system_prompt = """You are a financial analyst assistant. 
Analyze the user's spending based on the provided summary. 

A transaction table and a pie chart visualizing these categories are also shown to the user separately.

Generate a response strictly following this structure:

### Category Summary
[List each category, its total amount, and its percentage of total spending, like: 'Food: $44.37 (14.6%)']

### Observations
[Provide 2-3 bullet points observing spending patterns, like high spending areas or frequency.]

### Recommendations
[Provide 2-3 actionable, specific bullet-point recommendations for managing expenses or saving based *only* on the observed patterns.]

Maintain a helpful, non-judgmental tone. Be concise.
"""
        
        # Include context passed from app.py (period, total, count) 
        # and the calculated summary for the LLM to structure its response
        user_prompt = f"""
Spending Period: {period_str}
Total Expenses: {total_spent_str} across {num_transactions} transactions.

Category Breakdown:
{category_summary_text}

Please generate the structured financial analysis (Category Summary, Observations, Recommendations) based ONLY on this data.
"""
        # --- End New Prompt ---

        from langchain.schema import SystemMessage, HumanMessage
        response = llm.invoke([SystemMessage(content=system_prompt), HumanMessage(content=user_prompt)])
        analysis_text = response.content.strip()
        
        # Basic validation: Check if expected headers are present
        if "Category Summary" not in analysis_text or "Observations" not in analysis_text or "Recommendations" not in analysis_text:
             print("Warning: LLM output might not follow the requested structure.")
             print(f"LLM Raw Output: {analysis_text}")
             # Optionally prepend a note or return a fallback message
             analysis_text = "*LLM failed to generate fully structured analysis. Raw output below:*\n\n" + analysis_text
        else:
            print("Generated structured analysis successfully.")

        return analysis_text
        
    except Exception as e:
        print(f"Error generating structured financial analysis: {e}")
        import traceback
        traceback.print_exc()
        return "Error generating financial analysis report."
